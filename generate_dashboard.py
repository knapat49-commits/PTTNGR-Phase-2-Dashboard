"""
generate_dashboard.py  (v3 — with COST sheet support + GitHub Pages auto-push)
Watches the Phase 2 Excel file and regenerates the HTML dashboard on every save.
"""

import os, sys, time, json, datetime, subprocess
import html as _html
import urllib.request
import openpyxl

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE    = os.path.join(BASE_DIR, "Weekly Progress Phase2 update on 2026-06-17.xlsx")
HTML_OUT      = os.path.join(BASE_DIR, "Weekly_Progress_Phase2_Week5.html")
CHARTJS_LOCAL = os.path.join(BASE_DIR, "chart.umd.min.js")
CHARTJS_CDN   = "https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"
INTERVAL      = 5

# ── helpers ────────────────────────────────────────────────────────────────────

def p(v):    return f"{(v or 0)*100:.2f}%"
def pct0(v): return f"{(v or 0)*100:.0f}%"
def pv(v):   return round((v or 0)*100, 4)

def fdate(v):
    return v.strftime("%d-%b-%y") if isinstance(v, datetime.datetime) else (str(v) if v else "")

def perf_sign(v):
    if v is None: return "0.00%"
    return ("▲ +" if v >= 0 else "▼ ") + f"{v*100:.2f}%"

def perf_cls(v): return "perf-pos" if (v or 0) >= 0 else "perf-neg"
def esc(v):      return _html.escape(str(v)) if v is not None else ""

def pbar(v, cls):
    w = min(100, abs(pv(v)))
    if w == 0: return ""
    return ('<br><span class="pbar-outer"><span class="' + cls
            + '" style="width:' + str(w) + '%"></span></span>')

def fmt_money(v):
    if v is None: return '<span style="color:#999;">—</span>'
    s = f"{abs(v):,.0f}"
    if v < 0: return '<span style="color:#CC0000;">(' + s + ')</span>'
    return s

def fmt_pct_cost(v):
    if v is None: return '<span style="color:#999;">—</span>'
    s = f"{v*100:.0f}%"
    if v < 0: return '<span style="color:#CC0000;">' + s + '</span>'
    if v > 1: return '<span style="color:#CC0000;font-weight:bold;">' + s + '</span>'
    return s

# ── Chart.js cache ──────────────────────────────────────────────────────────────

def ensure_chartjs():
    if os.path.exists(CHARTJS_LOCAL): return True
    try:
        print("Downloading Chart.js ...", end=" ", flush=True)
        urllib.request.urlretrieve(CHARTJS_CDN, CHARTJS_LOCAL)
        print("done."); return True
    except Exception as e:
        print("\n  WARNING: download failed. Using CDN."); return False

# ── Excel reading ───────────────────────────────────────────────────────────────

def force_calculate(path):
    """Open the workbook in Excel (hidden), force-calculate all formulas, save.
    This populates the formula cache so openpyxl can read real values.
    Requires xlwings + Microsoft Excel installed (Windows only).
    Safe no-op on Linux/CI where xlwings/Excel is unavailable."""
    try:
        import xlwings as xw
        print("  [xlwings] Opening Excel to calculate formulas ...", end=" ", flush=True)
        app = xw.App(visible=False, add_book=False)
        try:
            wb = app.books.open(os.path.abspath(path))
            app.calculate()          # force full recalculation
            wb.save()                # write calculated values into cache
            wb.close()
            print("done.")
        finally:
            app.quit()
        return True
    except ImportError:
        return False   # xlwings not installed — CI/Linux, use existing cache
    except Exception as e:
        print(f"\n  [xlwings] Warning: {e}")
        return False


def read_excel(path):
    force_calculate(path)   # populates cache on Windows; safe no-op on Linux
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except PermissionError:
        raise PermissionError("Excel file is open. Close it, save, then wait.")
    missing = [s for s in ("CHART","DATA") if s not in wb.sheetnames]
    if missing:
        raise ValueError("Missing sheets: " + str(missing))
    result = {**_read_chart(wb["CHART"]), **_read_data(wb["DATA"])}
    result["cost"] = _read_cost(wb["COST"]) if "COST" in wb.sheetnames else None
    return result


def _read_chart(ws):
    def c(r, col): return ws.cell(r, col).value
    for row, col, expected, label in [
        (6,  2, "No.", "CHART!B6"),
        (14, 2, "No.", "CHART!B14"),
        (22, 2, "No.", "CHART!B22"),
    ]:
        got = c(row, col)
        if got != expected:
            raise ValueError(f"Layout mismatch at {label}: expected {expected!r}, got {got!r}")

    def prows(start, end):
        return [{"no":c(r,2),"wtg":c(r,3),"group":c(r,4),
                 "start":fdate(c(r,5)),"finish":fdate(c(r,6)),
                 "plan":c(r,7),"actual":c(r,8),"perf":c(r,9)} for r in range(start,end+1)]

    rspan={9:5,14:3,17:1,18:1}; rval={9:3,14:1,17:5,18:2}
    gspan={9:3,12:2,14:2,16:2,18:1}; gval={9:"B",12:"C",14:"D",16:"E",18:"F"}
    locs=[]
    for r in range(9,19):
        locs.append({"no":c(r,11),"region_span":rspan.get(r),"region_val":rval.get(r),
            "group_span":gspan.get(r),"group_val":gval.get(r),
            "tag":c(r,14),"rc":c(r,15),"location":c(r,16),
            "boring":c(r,17),"survey":c(r,18),"eng_doc":c(r,19),"permit":c(r,20),
            "excav":c(r,21),"inspect":c(r,22),"reinstate":c(r,23),"package":c(r,24),
            "prog_plan":c(r,25),"prog_actual":c(r,26),"is_grp_e":r in {16,17}})
    return {"week":c(1,2) or "Week ?","update_date":fdate(c(3,3)),
            "overall":prows(7,11),"overall_tot":{"plan":c(12,7),"actual":c(12,8),"perf":c(12,9)},
            "eng":prows(15,19),"eng_tot":{"plan":c(20,7),"actual":c(20,8),"perf":c(20,9)},
            "con":prows(23,27),"con_tot":{"plan":c(28,7),"actual":c(28,8),"perf":c(28,9)},
            "locations":locs,"accum_plan":c(19,25),"accum_actual":c(20,26)}


def _read_data(ws):
    def cv(r,col): return ws.cell(r,col).value
    tor,mplan,actual=[],[],[]; empty=0
    for col in range(8,89):
        t,m,a=cv(4,col),cv(5,col),cv(6,col)
        if t is None and m is None and a is None:
            empty+=1
            if empty>=3: break
            tor.append(None);mplan.append(None);actual.append(None)
        else:
            empty=0
            tor.append(round(t*100,4) if t is not None else None)
            mplan.append(round(m*100,4) if m is not None else None)
            actual.append(round(a*100,4) if a is not None else None)
    while tor and tor[-1] is None and mplan[-1] is None and actual[-1] is None:
        tor.pop();mplan.pop();actual.pop()
    bar=[{"group":cv(r,2),"plan":round((cv(r,3) or 0)*100,4),
          "actual":round((cv(r,4) or 0)*100,4),"perf":round((cv(r,5) or 0)*100,4)}
         for r in range(8,13)]
    return {"tor":tor,"mplan":mplan,"actual":actual,"bar":bar}


def _read_cost(ws):
    def num(r,col):
        v=ws.cell(r,col).value
        try: return float(v) if v not in (None,"") else None
        except: return None
    def txt(r,col):
        v=ws.cell(r,col).value
        return str(v) if v not in (None,"") else ""
    mhrs={"plan":num(10,2),"act_norm":num(11,2),"act_ot":num(12,2),
          "act_tot":num(13,2),"etc":num(14,2),"eac":num(15,2),"variance":num(16,2)}
    labels=["Revenue","MH Cost","Variable Cost","Material Cost","Total Direct Cost",
            "Contribution Margin","Allocation Cost","Total Cost (Direct+OH)",
            "Gross Margin","Total Cost (Inc. SG&A and DP)","Net Income"]
    rows=[{"label":lb,"plan":num(20+i,2),"act":num(20+i,3),"commit":num(20+i,4),
           "etc":num(20+i,5),"eac":num(20+i,6),"pct":num(20+i,7)}
          for i,lb in enumerate(labels)]
    return {"report_date":txt(5,2),"svo":txt(6,2),"week_label":txt(7,2),"mhrs":mhrs,"rows":rows}

# ── HTML: progress tables ────────────────────────────────────────────────────────

def progress_table(rows, total, title, actual_label):
    def rhtml(r):
        wtg = (str(int(round(r["wtg"]*100))) + "%" if isinstance(r["wtg"],(int,float))
               else esc(r["wtg"]))
        return ('\n      <tr class="data-row">'
                + '<td>' + esc(r["no"]) + '</td><td>' + wtg + '</td>'
                + '<td>' + esc(r["group"]) + '</td>'
                + '<td>' + esc(r["start"]) + '</td><td>' + esc(r["finish"]) + '</td>'
                + '<td>' + p(r["plan"])   + pbar(r["plan"],  "pbar-plan")   + '</td>'
                + '<td>' + p(r["actual"]) + pbar(r["actual"],"pbar-actual") + '</td>'
                + '<td class="' + perf_cls(r["perf"]) + '">' + perf_sign(r["perf"]) + '</td>'
                + '</tr>')
    t=total; rows_html="".join(rhtml(r) for r in rows)
    return ('\n    <table>\n      <colgroup>'
            '<col style="width:24px"><col style="width:32px"><col style="width:32px">'
            '<col style="width:60px"><col style="width:60px">'
            '<col style="width:52px"><col style="width:56px"><col style="width:60px">'
            '</colgroup>'
            '\n      <tr class="sec-hdr"><td colspan="3">' + esc(title) + '</td>'
            '<td colspan="2">Master Plan Baseline Rev.0</td>'
            '<td colspan="3">' + esc(actual_label) + '</td></tr>'
            '\n      <tr class="col-hdr"><td>No.</td><td>WTG.</td><td>Group</td>'
            '<td>Start</td><td>Finish</td><td>% Plan</td><td>% Actual</td><td>% Performance</td></tr>'
            + rows_html
            + '\n      <tr class="total-row"><td colspan="5">Total</td>'
            + '<td>' + p(t["plan"])   + pbar(t["plan"],  "pbar-plan")   + '</td>'
            + '<td>' + p(t["actual"]) + pbar(t["actual"],"pbar-actual") + '</td>'
            + '<td class="' + perf_cls(t["perf"]) + '">' + perf_sign(t["perf"]) + '</td></tr>'
            + '\n    </table>')


def location_table_html(locs, accum_plan, accum_actual):
    RC={3:"#FFC000",1:"#00B0F0",5:"#FFFF00",2:"#7030A0"}; RT={2:"color:#fff;"}
    rows=[]
    for loc in locs:
        rcell=gcell=""
        if loc["region_span"] is not None:
            rv=loc["region_val"]
            rcell=('<td rowspan="'+str(loc["region_span"])+'" style="background:'
                   +RC.get(rv,"")+';'+RT.get(rv,"")+' font-weight:bold;">'+str(rv)+'</td>')
        if loc["group_span"] is not None:
            cls=' class="grp-e-cell"' if loc["group_val"]=="E" else ""
            gcell='<td rowspan="'+str(loc["group_span"])+'"'+cls+'>'+esc(loc["group_val"])+'</td>'
        rc=' class="loc-data grp-e"' if loc["is_grp_e"] else ' class="loc-data"'
        rows.append(
            '\n      <tr'+rc+'>'
            +'<td>'+esc(loc["no"])+'</td>'+rcell+gcell
            +'<td>'+esc(loc["tag"])+'</td>'
            +'<td style="font-size:8px;">'+esc(loc["rc"])+'</td>'
            +'<td class="td-loc">'+esc(loc["location"])+'</td>'
            +'<td>'+pct0(loc["boring"])+'</td><td>'+pct0(loc["survey"])+'</td>'
            +'<td>'+pct0(loc["eng_doc"])+'</td><td>'+pct0(loc["permit"])+'</td>'
            +'<td>'+pct0(loc["excav"])+'</td><td>'+pct0(loc["inspect"])+'</td>'
            +'<td>'+pct0(loc["reinstate"])+'</td><td>'+pct0(loc["package"])+'</td>'
            +'<td>'+p(loc["prog_plan"])+pbar(loc["prog_plan"],"pbar-plan")+'</td>'
            +'<td>'+p(loc["prog_actual"])+pbar(loc["prog_actual"],"pbar-actual")+'</td>'
            +'</tr>')
    HDR=('\n    <table>\n      <colgroup>'
         '<col style="width:22px"><col style="width:36px"><col style="width:30px">'
         '<col style="width:36px"><col style="width:82px"><col style="width:138px">'
         '<col style="width:38px"><col style="width:38px"><col style="width:38px"><col style="width:38px">'
         '<col style="width:38px"><col style="width:38px"><col style="width:38px"><col style="width:38px">'
         '<col style="width:44px"><col style="width:50px"></colgroup>'
         '\n      <tr class="loc-hdr-main">'
         '<th rowspan="4">No</th><th rowspan="4">Region</th><th rowspan="4">Group</th>'
         '<th rowspan="4">Tag</th><th rowspan="4">RC</th><th rowspan="4">Location</th>'
         '<th colspan="4" style="background:#FFFF00;color:#000;">Engineering</th>'
         '<th colspan="4" style="background:#00B0F0;color:#000;">Construction</th>'
         '<th colspan="2" style="background:#fff;color:#000;">Progress</th></tr>'
         '\n      <tr>'
         '<th rowspan="2" style="background:#FFFF00;color:#000;font-weight:bold;font-size:8px;">Boring Log</th>'
         '<th rowspan="2" style="background:#FFFF00;color:#000;font-weight:bold;font-size:8px;">Survey &amp;<br>Verify</th>'
         '<th rowspan="2" style="background:#FFFF00;color:#000;font-weight:bold;font-size:8px;">Engineering<br>Document</th>'
         '<th rowspan="2" style="background:#FFFF00;color:#000;font-weight:bold;font-size:8px;">Area Permit</th>'
         '<th rowspan="2" style="background:#00B0F0;color:#000;font-weight:bold;font-size:8px;">Excavation</th>'
         '<th rowspan="2" style="background:#00B0F0;color:#000;font-weight:bold;font-size:8px;">Inspection</th>'
         '<th rowspan="2" style="background:#00B0F0;color:#000;font-weight:bold;font-size:8px;">Re-instate</th>'
         '<th rowspan="2" style="background:#00B0F0;color:#000;font-weight:bold;font-size:8px;">Package</th>'
         '<th rowspan="2" style="font-weight:bold;font-size:8px;">% Plan</th>'
         '<th rowspan="2" style="font-weight:bold;font-size:8px;">% Actual</th></tr>'
         '\n      <tr></tr>'
         '\n      <tr class="loc-hdr-wtg">'
         '<td class="wtg-eng">WTG. 15%</td><td class="wtg-eng">WTG. 25%</td>'
         '<td class="wtg-eng">WTG. 50%</td><td class="wtg-eng">WTG. 10%</td>'
         '<td class="wtg-con">WTG. 50%</td><td class="wtg-con">WTG. 30%</td>'
         '<td class="wtg-con">WTG. 10%</td><td class="wtg-con">WTG. 10%</td>'
         '<td style="font-weight:bold;">100%</td><td style="font-weight:bold;">100%</td></tr>')
    return (HDR+"".join(rows)
            +'\n      <tr class="accum-row"><td colspan="14" style="text-align:right;">Accumulate Plan %</td>'
            +'<td class="accum-val">'+p(accum_plan)+'</td><td></td></tr>'
            +'\n      <tr class="accum-row"><td colspan="14" style="text-align:right;">Accumulate Actual %</td>'
            +'<td></td><td class="accum-val">'+p(accum_actual)+'</td></tr>'
            +'\n    </table>')

# ── HTML: cost section ──────────────────────────────────────────────────────────

def cost_section_html(cost):
    if cost is None:
        return ('<div style="margin-top:14px;padding:12px;border:1px solid #bbb;border-radius:4px;'
                'background:#FFF8E1;font-size:11px;color:#888;">'
                '&#9888;&#65039;  No COST sheet found. '
                'Run <b>setup_cost_sheet.py</b> once, fill in the YELLOW cells, then save.</div>')
    m=cost["mhrs"]

    def mbox(label, val):
        disp=f"{val:,.0f}" if val is not None else "—"
        clr="#CC0000" if (val is not None and val<0) else "#222"
        return ('<div style="flex:1;min-width:90px;border:1px solid #ddd;border-radius:4px;'
                'padding:8px 4px;text-align:center;background:#fff;">'
                '<div style="font-size:18px;font-weight:bold;color:'+clr+';">'+disp+'</div>'
                '<div style="font-size:9px;color:#666;margin-top:2px;">'+label+'</div></div>')

    mhrs_html=('<div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px;">'
               +mbox("Plan MHRS",m["plan"])+mbox("Act Normal",m["act_norm"])
               +mbox("Act OT",m["act_ot"])+mbox("Total Act MHRS",m["act_tot"])
               +mbox("ETC Mhrs",m["etc"])+mbox("EAC MHRS",m["eac"])
               +mbox("Variance MHRS",m["variance"])+'</div>')

    hdrs=["Detail","Plan (BG)","Act Cost to Date","Commit","ETC Cost","EAC Cost","%"]
    hdr_row="".join('<th style="background:#D9D9D9;padding:4px 6px;font-size:10px;">'+h+'</th>' for h in hdrs)

    SUBTOTALS={"Total Direct Cost","Total Cost (Direct+OH)","Total Cost (Inc. SG&A and DP)","Net Income"}
    HIGHLIGHTS={"Total Cost (Inc. SG&A and DP)","Net Income"}
    rows_html=""
    for r in cost["rows"]:
        is_sub=r["label"] in SUBTOTALS; is_hi=r["label"] in HIGHLIGHTS
        bg="#FFF2CC" if is_hi else ("#F2F2F2" if is_sub else "#fff")
        fw="bold" if is_sub else "normal"
        indent="" if is_sub else "padding-left:14px;"
        rows_html+=('<tr style="background:'+bg+';">'
                    +'<td style="text-align:left;'+indent+'font-weight:'+fw+';padding:3px 6px;">'
                    +esc(r["label"])+'</td>'
                    +'<td>'+fmt_money(r["plan"])+'</td>'
                    +'<td>'+fmt_money(r["act"])+'</td>'
                    +'<td>'+fmt_money(r["commit"])+'</td>'
                    +'<td>'+fmt_money(r["etc"])+'</td>'
                    +'<td style="font-weight:'+fw+';">'+fmt_money(r["eac"])+'</td>'
                    +'<td>'+fmt_pct_cost(r["pct"])+'</td>'
                    +'</tr>')

    tbl=('<table style="width:100%;border-collapse:collapse;font-size:10px;">'
         '<colgroup><col style="width:32%"><col style="width:12%"><col style="width:14%">'
         '<col style="width:10%"><col style="width:12%"><col style="width:12%"><col style="width:8%"></colgroup>'
         '<tr>'+hdr_row+'</tr>'+rows_html+'</table>')

    return ('<div style="margin-top:14px;border:1px solid #bbb;border-radius:4px;padding:10px;background:#fff;">'
            '<div style="font-weight:bold;font-size:11px;text-align:center;background:#0070C0;color:#fff;'
            'padding:4px;border-radius:2px;margin-bottom:10px;">'
            'COST PERFORMANCE &nbsp;|&nbsp; SVO: '+esc(cost["svo"])
            +' &nbsp;|&nbsp; As of: '+esc(cost["report_date"])+'</div>'
            +mhrs_html+tbl+'</div>')

# ── HTML: full page ─────────────────────────────────────────────────────────────

def build_html(d, chartjs_src):
    week=esc(d["week"]); update_date=esc(d["update_date"])
    ts=datetime.datetime.now().strftime("%Y%m%d%H%M%S")

    n=len(d["tor"])
    actual_pad=d["actual"]+[None]*max(0,n-len(d["actual"]))
    mplan_pad=(d["mplan"]+[d["mplan"][-1]]*(n-len(d["mplan"]))
               if d["mplan"] and len(d["mplan"])<n else d["mplan"])

    tor_js=json.dumps(d["tor"]); mplan_js=json.dumps(mplan_pad)
    actual_js=json.dumps(actual_pad); labels_js=json.dumps(["Wk "+str(i) for i in range(n)])
    bar=d["bar"]
    bg_j=json.dumps(["rgba(112,173,71,0.40)" if b["perf"]>=0 else "rgba(255,80,80,0.45)" for b in bar])
    brd_j=json.dumps(["#375623" if b["perf"]>=0 else "#CC0000" for b in bar])
    bg_g_j=json.dumps([b["group"] for b in bar]); bp_j=json.dumps([b["plan"] for b in bar])
    ba_j=json.dumps([b["actual"] for b in bar]); bperf_j=json.dumps([b["perf"] for b in bar])

    left=(progress_table(d["overall"],d["overall_tot"],"OVERALL PROGRESS","OVERALL Actual Progress")
          +'\n    <div class="section-gap"></div>'
          +progress_table(d["eng"],d["eng_tot"],"ENGINEERING PROGRESS","ENGINEERING Actual Progress")
          +'\n    <div class="section-gap"></div>'
          +progress_table(d["con"],d["con_tot"],"CONSTRUCTION PROGRESS","CONSTRUCTION Actual Progress"))
    right=location_table_html(d["locations"],d["accum_plan"],d["accum_actual"])

    CSS=("* { box-sizing: border-box; margin: 0; padding: 0; }\n"
         "body { font-family: Arial, sans-serif; font-size: 10px; background: #fff; color: #000; padding: 10px; }\n"
         ".top-bar { display: flex; align-items: stretch; gap: 5px; margin-bottom: 6px; flex-wrap: wrap; }\n"
         ".week-badge { background: #FFFF00; border: 1px solid #888; font-weight: bold; font-size: 13px; padding: 4px 14px; display: flex; align-items: center; }\n"
         ".update-box { background: #FF66FF; border: 1px solid #888; font-weight: bold; padding: 4px 10px; display: flex; align-items: center; white-space: nowrap; }\n"
         ".group-status-header { background: #0070C0; color: #fff; font-weight: bold; font-size: 12px; padding: 4px 18px; border: 1px solid #888; display: flex; align-items: center; justify-content: center; }\n"
         ".location-status-header { background: #4472C4; color: #fff; font-weight: bold; font-size: 12px; padding: 4px 18px; border: 1px solid #888; flex: 1; display: flex; align-items: center; justify-content: center; }\n"
         ".main-layout { display: flex; gap: 6px; align-items: flex-start; }\n"
         ".left-panel { flex: 0 0 auto; width: 400px; }\n"
         ".right-panel { flex: 1; min-width: 0; }\n"
         "table { border-collapse: collapse; width: 100%; table-layout: fixed; }\n"
         "td, th { border: 1px solid #777; padding: 2px 3px; vertical-align: middle; text-align: center; overflow: hidden; }\n"
         ".sec-hdr td { background: #CDE5FF; font-weight: bold; font-size: 9.5px; }\n"
         ".col-hdr td { background: #CDE5FF; font-weight: bold; font-size: 9px; }\n"
         ".total-row td { background: #FFCC00; font-weight: bold; }\n"
         ".data-row td { font-size: 9px; }\n"
         ".perf-pos { color: #006600; } .perf-neg { color: #CC0000; }\n"
         ".section-gap { height: 5px; }\n"
         ".pbar-outer { width: 100%; background: #ddd; border-radius: 2px; height: 7px; display: block; margin-top: 1px; }\n"
         ".pbar-plan   { height: 7px; border-radius: 2px; background: #5B9BD5; display: block; }\n"
         ".pbar-actual { height: 7px; border-radius: 2px; background: #70AD47; display: block; }\n"
         ".right-panel table { font-size: 9px; }\n"
         ".loc-hdr-main th { background: #000; color: #fff; font-weight: bold; text-align: center; font-size: 9px; white-space: normal; word-break: break-word; }\n"
         ".loc-hdr-wtg td { font-weight: bold; text-align: center; font-size: 8px; }\n"
         ".loc-hdr-wtg .wtg-eng { background: #FFFF00; } .loc-hdr-wtg .wtg-con { background: #00B0F0; }\n"
         ".loc-data td { text-align: center; font-size: 9px; }\n"
         ".loc-data .td-loc { text-align: left; white-space: normal; word-break: break-word; }\n"
         ".grp-e td { background: #CCFFFF; } .grp-e-cell { background: #CCFFFF; }\n"
         ".accum-row td { font-weight: bold; text-align: right; font-size: 9px; background: #F2F2F2; }\n"
         ".accum-row .accum-val { text-align: center; color: #0070C0; font-weight: bold; }")

    AUTO_JS=("var _TS="+json.dumps(ts)+";\n"
             "setTimeout(function _chk(){\n"
             "  fetch(location.href,{cache:'no-store'})\n"
             "    .then(function(r){return r.text();})\n"
             "    .then(function(t){\n"
             "      var m=t.match(/var _TS=(\"\\d+\")/);\n"
             "      if(m&&m[1]!==JSON.stringify(_TS)){location.reload();}\n"
             "      else{setTimeout(_chk,8000);}\n"
             "    }).catch(function(){setTimeout(_chk,8000);});\n"
             "},8000);\n")

    CHART_JS=("var lbl="+labels_js+",tor="+tor_js+",mp="+mplan_js+",act="+actual_js+";\n"
              "new Chart(document.getElementById('sc'),{type:'line',data:{labels:lbl,datasets:[\n"
              "  {label:'TOR',data:tor,borderColor:'#A9D18E',backgroundColor:'transparent',borderWidth:2,borderDash:[6,3],pointRadius:0,tension:0.4},\n"
              "  {label:'Master Plan',data:mp,borderColor:'#5B9BD5',backgroundColor:'rgba(91,155,213,0.08)',borderWidth:2.5,pointRadius:0,tension:0.4},\n"
              "  {label:'Actual',data:act,borderColor:'#FF6600',backgroundColor:'rgba(255,102,0,0.12)',borderWidth:2.5,pointRadius:4,pointBackgroundColor:'#FF6600',spanGaps:false,tension:0.4}\n"
              "]},options:{responsive:true,maintainAspectRatio:false,\n"
              "  interaction:{mode:'index',intersect:false},\n"
              "  plugins:{legend:{position:'top',labels:{font:{family:'Arial',size:10},boxWidth:16}},\n"
              "    tooltip:{callbacks:{label:function(c){return c.parsed.y!==null?c.dataset.label+': '+c.parsed.y.toFixed(2)+'%':null;}}}},\n"
              "  scales:{x:{ticks:{font:{family:'Arial',size:8},maxRotation:45},grid:{color:'rgba(0,0,0,0.06)'}},\n"
              "    y:{min:0,max:100,ticks:{font:{family:'Arial',size:9},callback:function(v){return v+'%';},stepSize:10},\n"
              "      grid:{color:'rgba(0,0,0,0.06)'},title:{display:true,text:'Cumulative Progress (%)',font:{size:9}}}}}\n"
              "});\n"
              "var bg="+bg_g_j+",bp="+bp_j+",ba="+ba_j+",bperf="+bperf_j+",bbg="+bg_j+",bbd="+brd_j+";\n"
              "new Chart(document.getElementById('bc'),{type:'bar',data:{labels:bg,datasets:[\n"
              "  {label:'% Plan',data:bp,backgroundColor:'#5B9BD5',borderColor:'#2F75B6',borderWidth:1},\n"
              "  {label:'% Actual',data:ba,backgroundColor:'#70AD47',borderColor:'#375623',borderWidth:1},\n"
              "  {label:'% Performance',data:bperf,backgroundColor:bbg,borderColor:bbd,borderWidth:1}\n"
              "]},options:{responsive:true,maintainAspectRatio:false,\n"
              "  plugins:{legend:{position:'top',labels:{font:{family:'Arial',size:10},boxWidth:14}},\n"
              "    tooltip:{callbacks:{label:function(c){return c.dataset.label+': '+(c.parsed.y>=0?'+':'')+c.parsed.y.toFixed(2)+'%';}}}},\n"
              "  scales:{x:{ticks:{font:{family:'Arial',size:10}},grid:{display:false}},\n"
              "    y:{ticks:{font:{family:'Arial',size:9},callback:function(v){return v+'%';}},\n"
              "      grid:{color:'rgba(0,0,0,0.07)'},title:{display:true,text:'Progress (%)',font:{size:9}}}}}\n"
              "});\n")

    return ("<!DOCTYPE html>\n<html lang=\"en\">\n<head>\n"
            "<meta charset=\"UTF-8\">\n"
            "<meta name=\"viewport\" content=\"width=device-width,initial-scale=1.0\">\n"
            "<title>Weekly Progress Phase 2 - "+week+" ("+update_date+")</title>\n"
            "<style>\n"+CSS+"\n</style>\n"
            "<script>\n"+AUTO_JS+"</script>\n"
            "</head>\n<body>\n"
            "<div class=\"top-bar\">\n"
            "  <div class=\"week-badge\">"+week+"</div>\n"
            "  <div class=\"update-box\">Update :&nbsp; "+update_date+"</div>\n"
            "  <div class=\"group-status-header\">GROUP WEEKLY STATUS</div>\n"
            "  <div class=\"update-box\">Update :&nbsp; "+update_date+"</div>\n"
            "  <div class=\"location-status-header\">LOCATION WEEKLY STATUS</div>\n"
            "</div>\n"
            "<div class=\"main-layout\">\n"
            "  <div class=\"left-panel\">"+left+"</div>\n"
            "  <div class=\"right-panel\">"+right+"</div>\n"
            "</div>\n"
            "<div style=\"margin-top:14px;display:flex;gap:12px;align-items:flex-start;flex-wrap:wrap;\">\n"
            "  <div style=\"flex:1;min-width:540px;border:1px solid #bbb;border-radius:4px;padding:10px;background:#fff;\">\n"
            "    <div style=\"font-weight:bold;font-size:11px;text-align:center;background:#0070C0;color:#fff;padding:4px;border-radius:2px;margin-bottom:8px;\">Overall Weekly Project Progress</div>\n"
            "    <div style=\"position:relative;height:260px;\"><canvas id=\"sc\"></canvas></div>\n"
            "  </div>\n"
            "  <div style=\"flex:0 0 380px;border:1px solid #bbb;border-radius:4px;padding:10px;background:#fff;\">\n"
            "    <div style=\"font-weight:bold;font-size:11px;text-align:center;background:#0070C0;color:#fff;padding:4px;border-radius:2px;margin-bottom:8px;\">Group Weekly Status</div>\n"
            "    <div style=\"position:relative;height:260px;\"><canvas id=\"bc\"></canvas></div>\n"
            "  </div>\n"
            "</div>\n"
            "<script src=\""+chartjs_src+"\"></script>\n"
            "<script>\n"+CHART_JS+"</script>\n"
            +cost_section_html(d.get("cost"))
            +"\n</body>\n</html>")

# ── GitHub auto-push ────────────────────────────────────────────────────────────

def git_push(html_path):
    repo_dir=os.path.dirname(html_path); fname=os.path.basename(html_path)
    ts=datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    chk=subprocess.run(["git","rev-parse","--git-dir"],cwd=repo_dir,capture_output=True,text=True)
    if chk.returncode!=0:
        print("  [git] Not a git repo - skipping (run 'git init' first)"); return
    try:
        subprocess.run(["git","add",fname],cwd=repo_dir,check=True,capture_output=True)
        diff=subprocess.run(["git","diff","--cached","--quiet"],cwd=repo_dir)
        if diff.returncode==0: print("  [git] No changes to push"); return
        subprocess.run(["git","commit","-m","dashboard update "+ts],cwd=repo_dir,check=True,capture_output=True)
        subprocess.run(["git","push"],cwd=repo_dir,check=True,capture_output=True)
        print("  [git] Pushed to GitHub Pages OK")
    except FileNotFoundError:
        print("  [git] git not installed - download from git-scm.com")
    except subprocess.CalledProcessError as e:
        print("  [git] Push failed: "+(e.stderr.strip() if e.stderr else str(e)))

# ── runner ──────────────────────────────────────────────────────────────────────

def generate(path, chartjs_src):
    ts=datetime.datetime.now().strftime("%H:%M:%S")
    print("["+ts+"] Reading "+os.path.basename(path)+" ...", end=" ", flush=True)
    try:
        d=read_excel(path)
        with open(HTML_OUT,"w",encoding="utf-8") as f:
            f.write(build_html(d,chartjs_src))
        print("-> Saved  ("+d["week"]+", "+d["update_date"]+")")
        git_push(HTML_OUT)
        return True
    except PermissionError as e:
        print("\n  [locked] "+str(e)); return False
    except ValueError as e:
        print("\n  [data error] "+str(e)); return False
    except Exception as e:
        print("\n  [error] "+type(e).__name__+": "+str(e)); return False


def watch(path):
    if not os.path.exists(path):
        print("ERROR: not found: "+path); sys.exit(1)
    ok=ensure_chartjs()
    src=os.path.basename(CHARTJS_LOCAL) if (ok and os.path.exists(CHARTJS_LOCAL)) else CHARTJS_CDN
    print("Watching : "+path+"\nOutput   : "+HTML_OUT+"\nInterval : "+str(INTERVAL)+"s | Ctrl+C to stop\n")
    last=None
    try:
        while True:
            try:
                mt=os.path.getmtime(path)
                if mt!=last:
                    if generate(path,src):
                        # Re-read mtime after generate (xlwings may re-save the file)
                        last=os.path.getmtime(path)
                    else:
                        print("  Will retry on next file change ...")
            except Exception as e:
                print("  Watch error: "+str(e))
            time.sleep(INTERVAL)
    except KeyboardInterrupt:
        print("\nStopped.")


if __name__=="__main__":
    if "--once" in sys.argv:
        # CI mode: generate once using CDN and exit
        args = [a for a in sys.argv[1:] if a != "--once"]
        path = args[0] if args else EXCEL_FILE
        ok = generate(path, CHARTJS_CDN)
        sys.exit(0 if ok else 1)
    else:
        watch(sys.argv[1] if len(sys.argv)>1 else EXCEL_FILE)

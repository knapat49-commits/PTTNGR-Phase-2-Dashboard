"""
setup_cost_sheet.py
===================
Run this script ONCE to add a COST sheet to your Excel file.

After running:
  1. Open the Excel file
  2. Go to the COST sheet
  3. Fill in the YELLOW cells every week (copy numbers from the PDF)
  4. Save — the dashboard updates automatically

Usage:  python setup_cost_sheet.py
"""

import os
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "Weekly Progress Phase2 update on 2026-06-17.xlsx")

# ── colours ──────────────────────────────────────────────────────────────────
YELLOW  = PatternFill("solid", fgColor="FFFF00")   # user fills these
BLUE_H  = PatternFill("solid", fgColor="0070C0")   # section header
GREY_H  = PatternFill("solid", fgColor="D9D9D9")   # column header
GREEN   = PatternFill("solid", fgColor="E2EFDA")   # read-only hint
WHITE   = PatternFill("solid", fgColor="FFFFFF")

WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD       = Font(name="Arial", bold=True, size=10)
NORMAL     = Font(name="Arial", size=10)
RED_BOLD   = Font(name="Arial", bold=True, color="CC0000", size=10)

CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT   = Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value="", fill=None, font=None,
             align=None, border=True, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:   cell.fill   = fill
    if font:   cell.font   = font or NORMAL
    if align:  cell.alignment = align
    if border: cell.border = thin_border()
    if num_fmt: cell.number_format = num_fmt
    return cell

# ── main ─────────────────────────────────────────────────────────────────────

def create_cost_sheet(path):
    print(f"Opening: {path}")
    wb = openpyxl.load_workbook(path)

    # Remove old COST sheet if it exists
    if "COST" in wb.sheetnames:
        del wb["COST"]
        print("  Replaced existing COST sheet.")

    ws = wb.create_sheet("COST", 0)   # insert as first sheet
    print("  Created COST sheet.")

    # ── column widths ──────────────────────────────────────────────────────
    col_widths = [32, 16, 18, 14, 14, 18, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 40

    # ══════════════════════════════════════════════════════════════════════
    # ROW 1 — Title
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A1:G1")
    set_cell(ws, 1, 1, "COST DATA  —  Performance by Job",
             fill=BLUE_H, font=Font(name="Arial", bold=True, size=14, color="FFFFFF"),
             align=CENTER, border=False)

    # ROW 2 — Instruction banner
    ws.merge_cells("A2:G2")
    set_cell(ws, 2, 1,
             "📋  HOW TO USE:  Every week, open this sheet and fill in the YELLOW cells "
             "by copying numbers from the 'Performance by Job' PDF.  "
             "Then save the file — the dashboard updates automatically.",
             fill=PatternFill("solid", fgColor="FFF2CC"),
             font=Font(name="Arial", size=10, bold=True),
             align=LEFT, border=False)

    # ROW 3 — spacer
    ws.row_dimensions[3].height = 6

    # ══════════════════════════════════════════════════════════════════════
    # ROWS 4-5 — Report info
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A4:G4")
    set_cell(ws, 4, 1, "REPORT INFO",
             fill=BLUE_H, font=WHITE_BOLD, align=CENTER, border=False)

    info = [
        (5, "As of Date  (e.g. 17-Jun-26)",  None,          "DD-MMM-YY"),
        (6, "SVO Number",                     "201019973010", None),
        (7, "Week",                            "Week 5",      None),
    ]
    for row, label, example, fmt in info:
        set_cell(ws, row, 1, label,   fill=WHITE, font=BOLD,   align=LEFT)
        c = set_cell(ws, row, 2, example or "", fill=YELLOW, font=NORMAL, align=CENTER)
        if fmt: c.number_format = fmt
        # merge cols 3-7 as empty (keep border)
        for col in range(3, 8):
            set_cell(ws, row, col, "", fill=WHITE, font=NORMAL)

    ws.row_dimensions[8].height = 6  # spacer

    # ══════════════════════════════════════════════════════════════════════
    # ROWS 9-16 — MANHOURS
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A9:G9")
    set_cell(ws, 9, 1, "MANHOURS (MHRS)",
             fill=BLUE_H, font=WHITE_BOLD, align=CENTER, border=False)

    mhrs_labels = [
        (10, "Plan MHRS",         380),
        (11, "Act Normal Time",    38),
        (12, "Act Over Time",       0),
        (13, "Total Act MHRS",     38),
        (14, "ETC Mhrs",        2567),
        (15, "EAC MHRS",        2605),
        (16, "Variance MHRS",  -2225),
    ]
    for row, label, example in mhrs_labels:
        set_cell(ws, row, 1, label,   fill=WHITE, font=BOLD,   align=LEFT)
        set_cell(ws, row, 2, example, fill=YELLOW, font=NORMAL, align=CENTER,
                 num_fmt="#,##0")
        for col in range(3, 8):
            set_cell(ws, row, col, "", fill=WHITE, font=NORMAL)

    ws.row_dimensions[17].height = 6  # spacer

    # ══════════════════════════════════════════════════════════════════════
    # ROWS 18-29 — COST TABLE
    # ══════════════════════════════════════════════════════════════════════
    ws.merge_cells("A18:G18")
    set_cell(ws, 18, 1, "COST  (THB)",
             fill=BLUE_H, font=WHITE_BOLD, align=CENTER, border=False)

    # Column headers row 19
    ws.row_dimensions[19].height = 36
    col_hdrs = ["Detail", "Plan (BG)", "Act Cost to Date (A)",
                "Commit (B)", "ETC Cost (C)", "EAC Cost (A+B+C)", "%"]
    for col, hdr in enumerate(col_hdrs, 1):
        set_cell(ws, 19, col, hdr,
                 fill=GREY_H, font=BOLD, align=CENTER)

    # Cost rows  (label, plan, act, commit, etc, eac, pct)
    cost_rows = [
        ("Revenue",                          1050000,  437500,      0,  612500, 1050000,   None),
        ("MH Cost",                           129040,   11139,      0,  105742,  116881,   0.11),
        ("Variable Cost",                     525000,  526050,      0,       0,  526050,   0.50),
        ("  └ Material Cost",                 525000,  526050,      0,       0,  526050,   None),
        ("Total Direct Cost",                 654040,  537189,      0,  105742,  642931,   0.61),
        ("Contribution Margin",               395960,  -99689,      0,       0,  407069,   0.39),
        ("Allocation Cost",                        0,  191360,      0,       0,  375316,   0.36),
        ("Total Cost (Direct + OH)",               0,  728548,      0,       0, 1018246,   0.97),
        ("Gross Margin",                           0, -291048,      0,       0,   31754,   0.03),
        ("Total Cost (Inc. SG&A and DP)",    1039924,  821755,      0,       0, 1154446,   1.10),
        ("Net Income",                          8061,       0,      0,       0, -104446,  -0.10),
    ]

    MONEY_FMT = '#,##0'
    PCT_FMT   = '0%'

    for i, (label, plan, act, commit, etc, eac, pct) in enumerate(cost_rows):
        row = 20 + i
        ws.row_dimensions[row].height = 18

        is_total = label.startswith("Total") or label == "Net Income"
        f_label = Font(name="Arial", bold=is_total, size=10)
        indent  = LEFT if not label.startswith("  ") else Alignment(
                      horizontal="left", vertical="center", indent=2)

        set_cell(ws, row, 1, label.strip(), fill=WHITE, font=f_label, align=indent)

        for col, val, fmt in [
            (2, plan,   MONEY_FMT),
            (3, act,    MONEY_FMT),
            (4, commit, MONEY_FMT),
            (5, etc,    MONEY_FMT),
            (6, eac,    MONEY_FMT),
        ]:
            cell = set_cell(ws, row, col, val if val != 0 else None,
                            fill=YELLOW, font=NORMAL, align=RIGHT, num_fmt=fmt)
            # Negative values → red
            if val is not None and val < 0:
                cell.font = Font(name="Arial", size=10, color="CC0000")

        pct_cell = set_cell(ws, row, 7, pct, fill=YELLOW, font=NORMAL,
                            align=CENTER, num_fmt=PCT_FMT)
        if pct is not None and pct < 0:
            pct_cell.font = Font(name="Arial", size=10, color="CC0000")

    # ══════════════════════════════════════════════════════════════════════
    # Bottom legend
    # ══════════════════════════════════════════════════════════════════════
    last_row = 20 + len(cost_rows) + 1
    ws.row_dimensions[last_row].height = 6
    legend_row = last_row + 1
    ws.merge_cells(f"A{legend_row}:G{legend_row}")
    set_cell(ws, legend_row, 1,
             "✏️  Fill in YELLOW cells only.  Do NOT add/remove rows or rename this sheet.",
             fill=PatternFill("solid", fgColor="FFF2CC"),
             font=Font(name="Arial", bold=True, size=9),
             align=LEFT, border=False)

    # Freeze panes below header
    ws.freeze_panes = "B5"

    wb.save(path)
    print(f"\n  Done! COST sheet added to:\n  {path}")
    print("\nNext steps:")
    print("  1. Open the Excel file")
    print("  2. Go to the COST sheet")
    print("  3. Update the YELLOW cells with data from this week's PDF")
    print("  4. Save — the dashboard will update automatically")


if __name__ == "__main__":
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: Excel file not found:\n  {EXCEL_FILE}")
    else:
        create_cost_sheet(EXCEL_FILE)
